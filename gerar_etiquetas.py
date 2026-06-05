#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gerador de etiquetas de rolo PILAR.

Lê o processo via stdin (JSON) e escreve o .xlsx em stdout (binário).

Entrada (stdin), aceita ambos os formatos:
  { "processo": { "numero", "cliente", "itens": [...] }, "logoPath": null }
  ou diretamente:
  { "numero", "cliente", "itens": [{ "codigo", "composicao", "largura_cm", "gsm" }] }

Toda mensagem de diagnóstico vai para stderr — stdout carrega apenas o binário .xlsx.
"""
import sys
import io
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

LOGO_PADRAO = '/opt/pilar-calculadora/logo_pilar.png'
IMPORTADOR  = 'PILAR IMPORTS LTDA'
CNPJ        = '43.954.200/0001-96'
PAIS        = 'CHINA'

THIN   = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL   = PatternFill('solid', fgColor='F5F5F5')
ALIGN  = Alignment(vertical='center', wrap_text=True)

FONT_NORMAL = Font(name='Arial', size=9)
FONT_LABEL  = Font(name='Arial', size=9, bold=True)
FONT_TURQ   = Font(name='Arial', size=9, bold=True, color='00B5AD')

# Fontes/alinhamentos da etiqueta (conforme o modelo: Calibri)
F_ET_LABEL = Font(name='Calibri', size=11, bold=True)
F_ET_VALUE = Font(name='Calibri', size=11)
F_ET_COMP  = Font(name='Calibri', size=9)
AL_CENTER  = Alignment(horizontal='center', vertical='center')
AL_COMP    = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def label_rows(proc, item, cliente):
    """As 10 linhas da etiqueta (label, valor, label2, valor2) — conforme o modelo.
    Sem GRAMATURA; linha LARGURA traz 'QNTY ROLOS' como campo duplo."""
    largura  = item.get('largura_cm')
    larg_txt = f'{largura}CM' if largura not in (None, '') else ''
    return [
        ('ORDEM NRO',      str(proc.get('numero', '') or ''),      None,         None),
        ('CODIGO',         str(item.get('codigo', '') or ''),      None,         None),
        ('COMPOSIÇÃO',     str(item.get('composicao', '') or ''),  None,         None),
        (' LARGURA',       larg_txt,                               'QNTY ROLOS', None),
        ('ROLO',           None,                                   'de',         None),
        ('CLIENTE',        str(cliente or ''),                     None,         None),
        ('PESO LÍQUIDO',   None,                                   'PESO BRUTO', None),
        ('IMPORTADOR',     IMPORTADOR,                             None,         None),
        ('CNPJ',           CNPJ,                                   None,         None),
        ('PAÍS DE ORIGEM', PAIS,                                   None,         None),
    ]


def sanitize_sheet_name(name, usados):
    base = re.sub(r'[:\\/?*\[\]]', '-', name)[:31] or 'ITEM'
    nome = base
    c = 1
    while nome in usados:
        c += 1
        nome = (base[:28] + '_' + str(c))
    usados.add(nome)
    return nome


def style_cell(ws, r, c, value, is_label, is_turq=False):
    cell = ws.cell(row=r, column=c, value=value if value != '' else None)
    cell.border = BORDER
    cell.alignment = ALIGN
    if is_label:
        cell.font = FONT_LABEL
        cell.fill = FILL
    elif is_turq:
        cell.font = FONT_TURQ
    else:
        cell.font = FONT_NORMAL
    return cell


def escrever_etiqueta(ws, R, base, spec, logo_factory, logo_cols):
    """Escreve uma etiqueta (10 linhas) a partir da linha R, com colunas:
    base=label, base+1=valor, base+2=label2, base+3=valor2 (merge até base+4).
    Logo fica nas 3 linhas acima (R-3..R-1); logo_cols=(from_col, to_col) 0-index."""
    lab, val, lab2, val2, mend = base, base + 1, base + 2, base + 3, base + 4

    # bordas em todo o retângulo da etiqueta (antes de mesclar)
    for i in range(10):
        for c in range(lab, mend + 1):
            cc = ws.cell(row=R + i, column=c)
            cc.border = BORDER
            cc.font = F_ET_VALUE
            cc.alignment = AL_CENTER

    for i, (l, v, l2, v2) in enumerate(spec):
        r = R + i
        lc = ws.cell(row=r, column=lab, value=l)
        lc.font = F_ET_LABEL; lc.alignment = AL_CENTER; lc.border = BORDER
        if l2 is None:
            # valor de largura total (merge valor..fim)
            ws.merge_cells(start_row=r, start_column=val, end_row=r, end_column=mend)
            vc = ws.cell(row=r, column=val, value=(v if v not in (None, '') else None))
            vc.border = BORDER
            if l == 'COMPOSIÇÃO':
                vc.font = F_ET_COMP; vc.alignment = AL_COMP
                ws.row_dimensions[r].height = 45
            else:
                vc.font = F_ET_VALUE; vc.alignment = AL_CENTER
        else:
            # linha com campo duplo: valor em col única + label2 + valor2 (merge val2..fim)
            vc = ws.cell(row=r, column=val, value=(v if v not in (None, '') else None))
            vc.font = F_ET_VALUE; vc.alignment = AL_CENTER; vc.border = BORDER
            l2c = ws.cell(row=r, column=lab2, value=l2)
            l2c.font = F_ET_LABEL; l2c.alignment = AL_CENTER; l2c.border = BORDER
            ws.merge_cells(start_row=r, start_column=val2, end_row=r, end_column=mend)
            v2c = ws.cell(row=r, column=val2, value=(v2 if v2 not in (None, '') else None))
            v2c.font = F_ET_VALUE; v2c.alignment = AL_CENTER; v2c.border = BORDER

    # Logo acima da etiqueta (mesclado nas 3 linhas); silencioso se ausente
    logo_top = R - 3
    ws.merge_cells(start_row=logo_top, start_column=lab, end_row=R - 1, end_column=mend)
    ws.row_dimensions[logo_top].height = 15
    ws.row_dimensions[logo_top + 1].height = 15
    ws.row_dimensions[R - 1].height = 31.5
    img = logo_factory()
    if img is not None:
        # OneCellAnchor com ext fixo (200x55px) — posiciona sem distorcer.
        # Âncora em logo_cols[0] (esquerda col=1/B, direita col=6/G), topo da área (R-4).
        logo_row = R - 4
        marker = AnchorMarker(col=logo_cols[0], colOff=pixels_to_EMU(10),
                              row=logo_row, rowOff=pixels_to_EMU(5))
        img.anchor = OneCellAnchor(_from=marker,
                                   ext=XDRPositiveSize2D(pixels_to_EMU(200), pixels_to_EMU(55)))
        ws.add_image(img)


def montar_qtye(wb, proc):
    """Aba QTYE PER CUSTOMERS: A=CLIENTE, B=CÓDIGO, C=COMPOSIÇÃO, D=LARGURA, E=QUANTIDADE."""
    ws = wb.create_sheet(title='QTYE PER CUSTOMERS')
    headers = ['CLIENTE', 'CÓDIGO', 'COMPOSIÇÃO', 'LARGURA', 'QUANTIDADE']
    for c, h in enumerate(headers, start=1):
        style_cell(ws, 1, c, h, is_label=True)
    ws.row_dimensions[1].height = 13.5

    cliente_proc = str(proc.get('cliente', '') or '')
    itens = proc.get('itens') or []

    def chave(it):
        return (str(it.get('cliente') or cliente_proc).lower(), str(it.get('codigo') or '').lower())

    for ri, item in enumerate(sorted(itens, key=chave), start=2):
        ws.row_dimensions[ri].height = 13.5
        cliente = str(item.get('cliente') or cliente_proc)
        largura = item.get('largura_cm')
        largura_txt = f'{largura}CM' if largura not in (None, '') else ''
        qtd = item.get('quantidade')
        style_cell(ws, ri, 1, cliente, is_label=False)
        style_cell(ws, ri, 2, str(item.get('codigo') or ''), is_label=False)
        style_cell(ws, ri, 3, str(item.get('composicao') or ''), is_label=False)
        style_cell(ws, ri, 4, largura_txt, is_label=False)
        style_cell(ws, ri, 5, qtd if qtd not in (None, '') else None, is_label=False)

    for col, w in {1: 24, 2: 22, 3: 42, 4: 12, 5: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def main():
    raw = sys.stdin.read()
    try:
        data = json.loads(raw) if raw.strip() else {}
    except Exception as e:
        sys.stderr.write('JSON inválido em stdin: %s\n' % e)
        sys.exit(1)

    proc = data.get('processo', data) if isinstance(data, dict) else {}
    logo_path = (data.get('logoPath') if isinstance(data, dict) else None) or LOGO_PADRAO
    itens = proc.get('itens') or []

    if not itens:
        sys.stderr.write('Processo sem itens.\n')
        sys.exit(1)

    # Fábrica de imagem: instância nova por uso (openpyxl não reaproveita a mesma). 80x30px.
    def logo_img_factory():
        if not logo_path or not os.path.isfile(logo_path):
            return None
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            # tamanho controlado pelo XDRPositiveSize2D no anchor (sem distorção)
            return img
        except Exception as e:
            sys.stderr.write('Logo ignorado (%s)\n' % e)
            return None

    cliente_proc = str(proc.get('cliente', '') or '').strip()
    nome_cliente = cliente_proc or str((itens[0].get('cliente') or '')).strip() or 'ETIQUETAS'

    wb = Workbook()
    wb.remove(wb.active)  # remove a planilha vazia inicial
    usados = set()
    ws = wb.create_sheet(title=sanitize_sheet_name(nome_cliente, usados))

    # larguras de coluna conforme o modelo (A margem; B-F esquerda; G separador; H-L direita)
    larguras = {1: 13, 2: 15.8, 3: 13, 4: 16.5, 5: 11.8, 6: 13.45,
                7: 3.36, 8: 15.8, 9: 13, 10: 16.5, 11: 11.8, 12: 13.45}
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # 2 etiquetas por linha; cada grupo ocupa 14 linhas (3 logo + 10 rótulos + 1 separação)
    for p in range(0, len(itens), 2):
        grupo = p // 2
        R = 5 + grupo * 14  # linha inicial dos rótulos
        esq = itens[p]
        cli_e = str(esq.get('cliente') or cliente_proc or '')
        escrever_etiqueta(ws, R, 2, label_rows(proc, esq, cli_e), logo_img_factory, (1, 5))
        if p + 1 < len(itens):
            dir_ = itens[p + 1]
            cli_d = str(dir_.get('cliente') or cliente_proc or '')
            escrever_etiqueta(ws, R, 8, label_rows(proc, dir_, cli_d), logo_img_factory, (6, 10))

    montar_qtye(wb, proc)

    buf = io.BytesIO()
    wb.save(buf)
    sys.stdout.buffer.write(buf.getvalue())


if __name__ == '__main__':
    main()
